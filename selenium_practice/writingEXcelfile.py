import openpyxl
from  openpyxl.styles import PatternFill 

file = r'C:\Users\91878\python_tutotial\selenium_practice\data1.xlsx'
workbook= openpyxl.load_workbook(file)
sheet = workbook.active


for r in range(1,7+1):
    print("  ")
    for c in range(1, 6+1):
        sheet.cell(r,c).value = "welcome"




workbook.save(file)  # save file after each execution





