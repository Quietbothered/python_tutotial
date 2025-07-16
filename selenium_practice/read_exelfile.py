import openpyxl
# file ---> workbook---->sheets--->rows---->cells


file = r"C:\Users\91878\python_tutotial\selenium_practice\data.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook['Sheet1']      # or we can use workbook.active for just 1 sheet in workbook


rows = sheet.max_row
cols = sheet.max_column

for r in range(1,rows+1):
    print("  ")
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value,end='  ')
        




